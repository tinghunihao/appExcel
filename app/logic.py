import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


ANHUI_COUNTY_CITY = {
    "瑶海区": "合肥市",
    "庐阳区": "合肥市",
    "蜀山区": "合肥市",
    "包河区": "合肥市",
    "长丰县": "合肥市",
    "肥东县": "合肥市",
    "肥西县": "合肥市",
    "庐江县": "合肥市",
    "巢湖市": "合肥市",
    "镜湖区": "芜湖市",
    "弋江区": "芜湖市",
    "鸠江区": "芜湖市",
    "湾沚区": "芜湖市",
    "繁昌区": "芜湖市",
    "南陵县": "芜湖市",
    "无为市": "芜湖市",
    "龙子湖区": "蚌埠市",
    "蚌山区": "蚌埠市",
    "禹会区": "蚌埠市",
    "淮上区": "蚌埠市",
    "怀远县": "蚌埠市",
    "五河县": "蚌埠市",
    "固镇县": "蚌埠市",
    "大通区": "淮南市",
    "田家庵区": "淮南市",
    "谢家集区": "淮南市",
    "八公山区": "淮南市",
    "潘集区": "淮南市",
    "凤台县": "淮南市",
    "寿县": "淮南市",
    "花山区": "马鞍山市",
    "雨山区": "马鞍山市",
    "博望区": "马鞍山市",
    "当涂县": "马鞍山市",
    "含山县": "马鞍山市",
    "和县": "马鞍山市",
    "杜集区": "淮北市",
    "相山区": "淮北市",
    "烈山区": "淮北市",
    "濉溪县": "淮北市",
    "铜官区": "铜陵市",
    "义安区": "铜陵市",
    "郊区": "铜陵市",
    "枞阳县": "铜陵市",
    "迎江区": "安庆市",
    "大观区": "安庆市",
    "宜秀区": "安庆市",
    "怀宁县": "安庆市",
    "太湖县": "安庆市",
    "宿松县": "安庆市",
    "望江县": "安庆市",
    "岳西县": "安庆市",
    "桐城市": "安庆市",
    "潜山市": "安庆市",
    "屯溪区": "黄山市",
    "黄山区": "黄山市",
    "徽州区": "黄山市",
    "歙县": "黄山市",
    "休宁县": "黄山市",
    "黟县": "黄山市",
    "祁门县": "黄山市",
    "琅琊区": "滁州市",
    "南谯区": "滁州市",
    "来安县": "滁州市",
    "全椒县": "滁州市",
    "定远县": "滁州市",
    "凤阳县": "滁州市",
    "天长市": "滁州市",
    "明光市": "滁州市",
    "颍州区": "阜阳市",
    "颍东区": "阜阳市",
    "颍泉区": "阜阳市",
    "临泉县": "阜阳市",
    "太和县": "阜阳市",
    "阜南县": "阜阳市",
    "颍上县": "阜阳市",
    "界首市": "阜阳市",
    "埇桥区": "宿州市",
    "砀山县": "宿州市",
    "萧县": "宿州市",
    "灵璧县": "宿州市",
    "泗县": "宿州市",
    "金安区": "六安市",
    "裕安区": "六安市",
    "叶集区": "六安市",
    "霍邱县": "六安市",
    "舒城县": "六安市",
    "金寨县": "六安市",
    "霍山县": "六安市",
    "谯城区": "亳州市",
    "涡阳县": "亳州市",
    "蒙城县": "亳州市",
    "利辛县": "亳州市",
    "贵池区": "池州市",
    "东至县": "池州市",
    "石台县": "池州市",
    "青阳县": "池州市",
    "宣州区": "宣城市",
    "郎溪县": "宣城市",
    "泾县": "宣城市",
    "绩溪县": "宣城市",
    "旌德县": "宣城市",
    "宁国市": "宣城市",
    "广德市": "宣城市",
}

ANHUI_CITIES = sorted(set(ANHUI_COUNTY_CITY.values()), key=len, reverse=True)


def normalize_text(value) -> str:
    return str(value or "").strip()


def normalize_phone(value) -> str:
    return re.sub(r"\D+", "", normalize_text(value))


def normalize_name(value) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def normalize_model(value) -> str:
    text = normalize_text(value).upper()
    text = text.replace("＋", "+").replace("－", "-").replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", "", text)


def find_anhui_city(address: str) -> Tuple[Optional[str], Optional[str]]:
    text = normalize_text(address)
    for county, city in sorted(ANHUI_COUNTY_CITY.items(), key=lambda item: len(item[0]), reverse=True):
        if county in text:
            return county, city
    for city in ANHUI_CITIES:
        if city in text:
            return None, city
    return None, None


def complete_anhui_address(address: str) -> Tuple[str, str]:
    text = normalize_text(address)
    if not text:
        return "", "空地址"

    county, city = find_anhui_city(text)
    if not city:
        if text.startswith("安徽省"):
            return text, "已含省份；未识别地市"
        return text, "未识别区县或地市，未补全"

    result = text
    if result.startswith("安徽省"):
        if city not in result[:12]:
            result = "安徽省" + city + result[3:]
            return result, f"已补地市：{city}"
        return result, "已完整"

    if result.startswith(city) or city in result[:10]:
        return "安徽省" + result, f"已补省份：安徽省"

    return "安徽省" + city + result, f"已补省市：安徽省{city}"


def header_map(ws) -> Dict[str, int]:
    headers = {}
    for cell in ws[1]:
        name = normalize_text(cell.value)
        if name and name not in headers:
            headers[name] = cell.column
    return headers


def find_header(headers: Dict[str, int], candidates: Iterable[str]) -> Optional[str]:
    for candidate in candidates:
        if candidate in headers:
            return candidate
    for candidate in candidates:
        for header in headers:
            if candidate in header:
                return header
    return None


def ensure_column(ws, headers: Dict[str, int], name: str) -> int:
    if name in headers:
        return headers[name]
    column = ws.max_column + 1
    ws.cell(row=1, column=column, value=name)
    headers[name] = column
    return column


def copy_row_style(ws, source_col: int, target_col: int) -> None:
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row=row, column=source_col)
        target = ws.cell(row=row, column=target_col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format


def normalize_code(value) -> str:
    text = normalize_text(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\s+", "", text)


def split_user_list(value: str) -> List[str]:
    parts = re.split(r"[,，;；\n\r]+", value or "")
    return [part.strip() for part in parts if part.strip()]


def parse_rename_map(value: str) -> Dict[str, str]:
    renames = {}
    for item in split_user_list(value):
        if "=" in item:
            old, new = item.split("=", 1)
        elif "：" in item:
            old, new = item.split("：", 1)
        elif ":" in item:
            old, new = item.split(":", 1)
        else:
            continue
        old = old.strip()
        new = new.strip()
        if old and new:
            renames[old] = new
    return renames


def parse_extract_columns(spec: str, headers: Dict[str, int]) -> List[Tuple[int, str]]:
    selected = []
    seen = set()
    max_column = max(headers.values(), default=0)
    column_names = {column: name for name, column in headers.items()}
    for item in split_user_list(spec):
        column = None
        source_name = None
        if item.isdigit():
            column = int(item)
            if column < 1:
                raise ValueError(f"列序号必须大于 0：{item}")
            if column > max_column:
                raise ValueError(f"列序号超出表头范围：{item}")
            source_name = column_names.get(column, f"第{column}列")
        elif item in headers:
            column = headers[item]
            source_name = item
        else:
            found = find_header(headers, [item])
            if found:
                column = headers[found]
                source_name = found
        if not column:
            raise ValueError(f"找不到要提取的列：{item}")
        if column not in seen:
            selected.append((column, source_name or item))
            seen.add(column)
    if not selected:
        raise ValueError("请填写要提取的列，例如：1,3,5,8,10")
    return selected


def build_master_lookup(master_ws, master_code_header: str) -> Dict[str, Dict[str, object]]:
    headers = header_map(master_ws)
    if master_code_header not in headers:
        raise ValueError(f"机型主数据表找不到列：{master_code_header}")
    merge_fields = ["编码", "系统机型", "进货价", "厂家机型"]
    missing = [field for field in merge_fields if field not in headers]
    if missing:
        raise ValueError("机型主数据表缺少列：" + "、".join(missing))

    lookup = {}
    for row in range(2, master_ws.max_row + 1):
        code = normalize_code(master_ws.cell(row=row, column=headers[master_code_header]).value)
        if not code or code in lookup:
            continue
        lookup[code] = {field: master_ws.cell(row=row, column=headers[field]).value for field in merge_fields}
    return lookup


def extract_and_merge_workbook(
    raw_path: Path,
    master_path: Path,
    output_path: Path,
    selected_columns_spec: str,
    rename_spec: str = "",
    raw_code_header: str = "69国标码",
    master_code_header: str = "69码",
) -> Dict[str, int]:
    raw_wb = load_workbook(raw_path, data_only=False)
    master_wb = load_workbook(master_path, data_only=False)
    raw_ws = raw_wb.active
    master_ws = master_wb.active
    raw_headers = header_map(raw_ws)

    if raw_code_header not in raw_headers:
        raise ValueError(f"原始表找不到匹配列：{raw_code_header}")

    selected = parse_extract_columns(selected_columns_spec, raw_headers)
    selected_columns = [col for col, _ in selected]
    if raw_headers[raw_code_header] not in selected_columns:
        selected.append((raw_headers[raw_code_header], raw_code_header))

    renames = parse_rename_map(rename_spec)
    master_lookup = build_master_lookup(master_ws, master_code_header)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "提取合并结果"

    merge_fields = ["编码", "系统机型", "进货价", "厂家机型"]
    base_headers = []
    for column, source_name in selected:
        new_name = renames.get(str(column), renames.get(source_name, source_name))
        base_headers.append(new_name)
    out_ws.append(base_headers + merge_fields + ["属性匹配状态"])

    code_selected_index = next(index for index, (column, _) in enumerate(selected) if column == raw_headers[raw_code_header])
    matched = 0
    not_found = 0
    for row in range(2, raw_ws.max_row + 1):
        values = [raw_ws.cell(row=row, column=column).value for column, _ in selected]
        code = normalize_code(values[code_selected_index])
        master = master_lookup.get(code)
        if master:
            out_ws.append(values + [master[field] for field in merge_fields] + ["成功"])
            matched += 1
        else:
            out_ws.append(values + ["", "", "", "", "未找到69码"])
            not_found += 1

    for column_cells in out_ws.columns:
        width = min(max(len(normalize_text(cell.value)) for cell in column_cells) + 2, 38)
        out_ws.column_dimensions[column_cells[0].column_letter].width = max(width, 10)

    out_wb.save(output_path)
    return {
        "rows": max(raw_ws.max_row - 1, 0),
        "matched": matched,
        "not_found": not_found,
        "columns": len(selected),
    }


def complete_address_workbook(input_path: Path, output_path: Path, address_column: str) -> Dict[str, int]:
    wb = load_workbook(input_path)
    ws = wb.active
    headers = header_map(ws)
    if address_column not in headers:
        raise ValueError(f"找不到地址列：{address_column}")

    address_col = headers[address_column]
    full_col = ensure_column(ws, headers, "补全地址")
    status_col = ensure_column(ws, headers, "地址补全状态")
    copy_row_style(ws, address_col, full_col)

    changed = 0
    unresolved = 0
    for row in range(2, ws.max_row + 1):
        original = ws.cell(row=row, column=address_col).value
        completed, status = complete_anhui_address(original)
        ws.cell(row=row, column=full_col, value=completed)
        ws.cell(row=row, column=status_col, value=status)
        if completed != normalize_text(original):
            changed += 1
        if status.startswith("未识别"):
            unresolved += 1

    wb.save(output_path)
    return {"rows": max(ws.max_row - 1, 0), "changed": changed, "unresolved": unresolved}


@dataclass
class MatchColumns:
    a_phone: str
    a_name: str
    a_model: str
    a_fuzzy_model: str
    b_phone: str
    b_name: str
    b_model: str
    b_ship_no: str


@dataclass
class BRecord:
    row: int
    phone: str
    name: str
    model: str
    raw_model: str
    ship_no: str


def infer_match_columns(a_headers: Dict[str, int], b_headers: Dict[str, int]) -> MatchColumns:
    a_phone = find_header(a_headers, ["收货人电话", "收货电话", "手机号", "手机号码", "电话"])
    a_name = find_header(a_headers, ["收货人姓名", "收货人", "姓名", "客户姓名"])
    a_model = find_header(a_headers, ["厂家机型", "机型", "产品信息", "商品名称", "产品名称", "型号"])
    a_fuzzy_model = find_header(a_headers, ["系统机型", "完整机型", "产品信息", "商品名称", "产品名称", "机型", "型号", "厂家机型"])
    b_phone = find_header(b_headers, ["收货人电话", "收货电话", "手机号", "手机号码", "电话"])
    b_name = find_header(b_headers, ["收货人姓名", "收货人", "姓名", "客户姓名"])
    b_model = find_header(b_headers, ["简版机型", "机型", "产品信息", "商品名称", "产品名称", "型号"])
    b_ship_no = find_header(b_headers, ["发货单号", "物流单号", "快递单号", "运单号", "单号"])
    missing = [
        label
        for label, value in [
            ("A表电话列", a_phone),
            ("A表姓名列", a_name),
            ("A表厂家机型列", a_model),
            ("A表模糊机型列", a_fuzzy_model),
            ("B表电话列", b_phone),
            ("B表姓名列", b_name),
            ("B表机型列", b_model),
            ("B表发货单号列", b_ship_no),
        ]
        if not value
    ]
    if missing:
        raise ValueError("自动识别列失败：" + "、".join(missing))
    return MatchColumns(a_phone, a_name, a_model, a_fuzzy_model, b_phone, b_name, b_model, b_ship_no)


def model_matches(a_model: str, b_model: str) -> bool:
    a = normalize_model(a_model)
    b = normalize_model(b_model)
    return bool(a and b and b in a)


def model_exact_matches(a_model: str, b_model: str) -> bool:
    a = normalize_model(a_model)
    b = normalize_model(b_model)
    return bool(a and b and a == b)


def model_fuzzy_matches(a_model: str, b_model: str) -> bool:
    a = normalize_model(a_model)
    b = normalize_model(b_model)
    return bool(a and b and (b in a or a in b))


def match_order_workbooks(
    a_path: Path,
    b_path: Path,
    output_path: Path,
    columns: Optional[MatchColumns] = None,
) -> Dict[str, int]:
    a_wb = load_workbook(a_path)
    b_wb = load_workbook(b_path, data_only=False)
    a_ws = a_wb.active
    b_ws = b_wb.active
    a_headers = header_map(a_ws)
    b_headers = header_map(b_ws)
    columns = columns or infer_match_columns(a_headers, b_headers)

    for required in columns.__dict__.values():
        if not required:
            raise ValueError("列名不能为空")
    for header in [columns.a_phone, columns.a_name, columns.a_model, columns.a_fuzzy_model]:
        if header not in a_headers:
            raise ValueError(f"A表找不到列：{header}")
    for header in [columns.b_phone, columns.b_name, columns.b_model, columns.b_ship_no]:
        if header not in b_headers:
            raise ValueError(f"B表找不到列：{header}")

    ship_col = ensure_column(a_ws, a_headers, "发货单号")
    status_col = ensure_column(a_ws, a_headers, "匹配状态")
    reason_col = ensure_column(a_ws, a_headers, "匹配原因")

    b_by_phone: Dict[str, List[BRecord]] = defaultdict(list)
    for row in range(2, b_ws.max_row + 1):
        phone = normalize_phone(b_ws.cell(row=row, column=b_headers[columns.b_phone]).value)
        if not phone:
            continue
        raw_model = normalize_text(b_ws.cell(row=row, column=b_headers[columns.b_model]).value)
        b_by_phone[phone].append(
            BRecord(
                row=row,
                phone=phone,
                name=normalize_name(b_ws.cell(row=row, column=b_headers[columns.b_name]).value),
                model=normalize_model(raw_model),
                raw_model=raw_model,
                ship_no=normalize_text(b_ws.cell(row=row, column=b_headers[columns.b_ship_no]).value),
            )
        )

    matched = 0
    not_found = 0
    ambiguous = 0
    for row in range(2, a_ws.max_row + 1):
        a_phone = normalize_phone(a_ws.cell(row=row, column=a_headers[columns.a_phone]).value)
        a_name = normalize_name(a_ws.cell(row=row, column=a_headers[columns.a_name]).value)
        a_model_raw = normalize_text(a_ws.cell(row=row, column=a_headers[columns.a_model]).value)
        a_fuzzy_model_raw = normalize_text(a_ws.cell(row=row, column=a_headers[columns.a_fuzzy_model]).value)

        if not a_phone:
            a_ws.cell(row=row, column=ship_col, value="单号未找到")
            a_ws.cell(row=row, column=status_col, value="失败")
            a_ws.cell(row=row, column=reason_col, value="A表电话为空")
            not_found += 1
            continue

        candidates = b_by_phone.get(a_phone, [])
        if not candidates:
            a_ws.cell(row=row, column=ship_col, value="单号未找到")
            a_ws.cell(row=row, column=status_col, value="失败")
            a_ws.cell(row=row, column=reason_col, value="电话不匹配：B表无相同电话")
            not_found += 1
            continue

        name_candidates = [item for item in candidates if item.name == a_name]
        if not name_candidates:
            a_ws.cell(row=row, column=ship_col, value="单号未找到")
            a_ws.cell(row=row, column=status_col, value="失败")
            a_ws.cell(row=row, column=reason_col, value="姓名不匹配：电话相同但收货人不同")
            not_found += 1
            continue

        exact_candidates = [item for item in name_candidates if model_exact_matches(a_model_raw, item.raw_model)]
        fuzzy_candidates = []
        if not exact_candidates:
            fuzzy_candidates = [item for item in name_candidates if model_fuzzy_matches(a_fuzzy_model_raw, item.raw_model)]

        model_candidates = exact_candidates or fuzzy_candidates
        match_method = "厂家机型精确匹配" if exact_candidates else "机型模糊匹配"
        if len(model_candidates) == 1:
            a_ws.cell(row=row, column=ship_col, value=model_candidates[0].ship_no)
            a_ws.cell(row=row, column=status_col, value="成功")
            a_ws.cell(
                row=row,
                column=reason_col,
                value=f"电话、姓名、{match_method}；B表第{model_candidates[0].row}行",
            )
            matched += 1
        elif len(model_candidates) > 1:
            a_ws.cell(row=row, column=ship_col, value="单号未找到")
            a_ws.cell(row=row, column=status_col, value="失败")
            a_ws.cell(row=row, column=reason_col, value="多条候选同时匹配，需人工确认")
            ambiguous += 1
        else:
            b_models = "；".join(item.raw_model for item in name_candidates[:5])
            a_ws.cell(row=row, column=ship_col, value="单号未找到")
            a_ws.cell(row=row, column=status_col, value="失败")
            a_ws.cell(row=row, column=reason_col, value=f"机型冲突：A表厂家机型[{a_model_raw}]，A表模糊机型[{a_fuzzy_model_raw}]，B表候选[{b_models}]")
            not_found += 1

    a_wb.save(output_path)
    return {
        "rows": max(a_ws.max_row - 1, 0),
        "matched": matched,
        "not_found": not_found,
        "ambiguous": ambiguous,
    }
