from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

from app.logic import (
    complete_anhui_address,
    complete_address_workbook,
    extract_and_merge_workbook,
    match_order_workbooks,
)


class LogicTests(unittest.TestCase):
    def test_complete_anhui_address(self):
        completed, status = complete_anhui_address("阜南县朱寨镇刘锋手机大卖场")
        self.assertEqual(completed, "安徽省阜阳市阜南县朱寨镇刘锋手机大卖场")
        self.assertIn("安徽省阜阳市", status)

        completed, _ = complete_anhui_address("阜阳市阜南县朱寨镇刘锋手机大卖场")
        self.assertEqual(completed, "安徽省阜阳市阜南县朱寨镇刘锋手机大卖场")

        completed, _ = complete_anhui_address("安徽省阜阳市阜南县朱寨镇刘锋手机大卖场")
        self.assertEqual(completed, "安徽省阜阳市阜南县朱寨镇刘锋手机大卖场")


    def test_workbook_address_and_match(self):
        with TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            address_input = tmp / "address.xlsx"
            address_output = tmp / "address_out.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["详细地址"])
            ws.append(["阜南县朱寨镇刘锋手机大卖场"])
            wb.save(address_input)

            stats = complete_address_workbook(address_input, address_output, "详细地址")
            self.assertEqual(stats["changed"], 1)
            out = load_workbook(address_output).active
            self.assertEqual(out["B2"].value, "安徽省阜阳市阜南县朱寨镇刘锋手机大卖场")

            a_file = tmp / "a.xlsx"
            b_file = tmp / "b.xlsx"
            result_file = tmp / "result.xlsx"

            a_wb = Workbook()
            a_ws = a_wb.active
            a_ws.append(["收货人电话", "收货人姓名", "厂家机型", "系统机型"])
            a_ws.append(["138-0000-0000", "张三", "vivo Y600 Pro 8＋512G", "错误的系统机型"])
            a_ws.append(["13900000000", "李四", "vivo X100 12G 256G", "vivo X100 12G 256G"])
            a_ws.append(["13700000000", "王五", "无法精确匹配", "vivo Y600 Pro 8＋512G(全网通)-浮光金"])
            a_wb.save(a_file)

            b_wb = Workbook()
            b_ws = b_wb.active
            b_ws.append(["收货人电话", "收货人姓名", "机型", "发货单号"])
            b_ws.append(["13800000000", "张三", "VIVO Y600 PRO 8+512G", "SF123"])
            b_ws.append(["13900000000", "李四", "vivo X90", "SF999"])
            b_ws.append(["13700000000", "王五", "vivo Y600 Pro 8+512G", "SF456"])
            b_wb.save(b_file)

            stats = match_order_workbooks(a_file, b_file, result_file)
            self.assertEqual(stats["matched"], 2)
            self.assertEqual(stats["not_found"], 1)

            result = load_workbook(result_file).active
            headers = {cell.value: cell.column for cell in result[1]}
            self.assertEqual(result.cell(2, headers["发货单号"]).value, "SF123")
            self.assertIn("厂家机型精确匹配", result.cell(2, headers["匹配原因"]).value)
            self.assertEqual(result.cell(3, headers["发货单号"]).value, "单号未找到")
            self.assertIn("机型冲突", result.cell(3, headers["匹配原因"]).value)
            self.assertEqual(result.cell(4, headers["发货单号"]).value, "SF456")
            self.assertIn("机型模糊匹配", result.cell(4, headers["匹配原因"]).value)

    def test_extract_and_merge_workbook(self):
        with TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            raw_file = tmp / "raw.xlsx"
            master_file = tmp / "master.xlsx"
            result_file = tmp / "extract_result.xlsx"

            raw_wb = Workbook()
            raw_ws = raw_wb.active
            raw_ws.append(["订单号", "客户", "69国标码", "数量", "备注"])
            raw_ws.append(["A001", "张三", "6901234567890", 1, "保留"])
            raw_ws.append(["A002", "李四", "6900000000000", 1, "未匹配"])
            raw_wb.save(raw_file)

            master_wb = Workbook()
            master_ws = master_wb.active
            master_ws.append(["69码", "编码", "系统机型", "进货价", "厂家机型"])
            master_ws.append(["6901234567890", "P001", "vivo Y600 Pro", 1599, "Y600 Pro 8+512G"])
            master_wb.save(master_file)

            stats = extract_and_merge_workbook(
                raw_file,
                master_file,
                result_file,
                "1,3,5",
                "1=订单编号,69国标码=商品条码",
            )
            self.assertEqual(stats["matched"], 1)
            self.assertEqual(stats["not_found"], 1)

            result = load_workbook(result_file).active
            headers = [cell.value for cell in result[1]]
            self.assertEqual(headers[:8], ["订单编号", "商品条码", "备注", "编码", "系统机型", "进货价", "厂家机型", "属性匹配状态"])
            self.assertEqual(result["D2"].value, "P001")
            self.assertEqual(result["E2"].value, "vivo Y600 Pro")
            self.assertEqual(result["H3"].value, "未找到69码")


if __name__ == "__main__":
    unittest.main()
